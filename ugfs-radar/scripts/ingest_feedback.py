"""
scripts/ingest_feedback.py — CLI standalone d'ingestion des décisions UGFS.

Alternative à l'endpoint POST /api/feedback/excel : ce script lit un
fichier Excel modifié en local et applique les décisions Go/No-Go en DB
sans passer par l'API. Utile pour :
  - tests d'intégration sans démarrer le service web
  - reprise manuelle si problème d'API

Usage :
    python -m scripts.ingest_feedback path/to/UGFS-Radar_2026-04-27.xlsx
    python -m scripts.ingest_feedback path/to/file.xlsx --by "alice@ugfs-na.com"
"""
from __future__ import annotations

import argparse
import asyncio
from pathlib import Path

from openpyxl import load_workbook

from src.config.logger import get_logger
from src.delivery.excel_builder import ALL_OPPS_COLUMNS
from src.storage.database import init_db, session_scope
from src.storage.repository import OpportunityRepo

logger = get_logger(__name__)

VALID_DECISIONS = {"GO", "NO_GO", "BORDERLINE", "SUBMITTED"}

COL_ID = ALL_OPPS_COLUMNS.index("ID")
COL_TITLE = ALL_OPPS_COLUMNS.index("Titre")
COL_DECISION = ALL_OPPS_COLUMNS.index("Décision interne (Go/No-Go)")
COL_REASON = ALL_OPPS_COLUMNS.index("Raison décision")


async def ingest(path: Path, submitted_by: str) -> dict:
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = "Toutes opportunités"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet '{sheet_name}' absent. Onglets: {wb.sheetnames}")
    ws = wb[sheet_name]

    n_processed, n_skipped, errors = 0, 0, []
    await init_db()

    async with session_scope() as session:
        repo = OpportunityRepo(session)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[COL_ID] is None:
                continue
            decision_raw = row[COL_DECISION]
            if not decision_raw:
                n_skipped += 1
                continue
            decision = str(decision_raw).strip().upper()
            if decision not in VALID_DECISIONS:
                errors.append(f"Ligne {i}: décision invalide '{decision}'")
                continue
            try:
                await repo.apply_feedback(
                    opportunity_id=int(row[COL_ID]),
                    decision=decision,
                    reason=str(row[COL_REASON]) if row[COL_REASON] else None,
                    submitted_by=submitted_by,
                )
                n_processed += 1
            except Exception as e:
                errors.append(f"Ligne {i}: {e}")

    logger.info("ingested", processed=n_processed, skipped=n_skipped, errors=len(errors))
    return {"processed": n_processed, "skipped": n_skipped, "errors": errors}


def main():
    p = argparse.ArgumentParser(description="Ingère les décisions Go/No-Go d'un Excel UGFS-Radar.")
    p.add_argument("path", type=Path, help="Chemin du fichier Excel modifié")
    p.add_argument("--by", default="cli", help="Auteur des décisions (pour audit trail)")
    args = p.parse_args()

    result = asyncio.run(ingest(args.path, args.by))
    print(f"\n✓ Ingestion terminée")
    print(f"  Décisions appliquées : {result['processed']}")
    print(f"  Lignes ignorées      : {result['skipped']}")
    if result["errors"]:
        print(f"  Erreurs              : {len(result['errors'])}")
        for e in result["errors"][:5]:
            print(f"    - {e}")


if __name__ == "__main__":
    main()
