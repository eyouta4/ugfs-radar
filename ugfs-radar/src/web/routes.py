"""
src/web/routes.py — Routes du dashboard web UGFS-Radar.

Pages
-----
  GET  /            → redirige vers /dashboard si connecté, sinon /login
  GET  /login       → page de login
  POST /login       → traitement formulaire login
  POST /logout      → suppression cookie
  GET  /dashboard   → tableau de bord (KPI + opportunités semaine)
  GET  /opportunity/{id} → fiche détaillée
  POST /api/decision → décision Go/No-Go (1 clic depuis dashboard)
  POST /api/upload-excel → upload du fichier Excel renvoyé par UGFS
  GET  /admin/users → gestion utilisateurs (admin only)
  POST /admin/users → créer un utilisateur (admin only)
  GET  /admin/runs  → historique des exécutions (admin only)
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Annotated

from fastapi import (
    APIRouter, Cookie, Depends, File, Form, HTTPException, Query,
    Request, Response, UploadFile, status,
)
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from sqlalchemy import desc, func, select

from src.config.logger import get_logger
from src.delivery.excel_builder import ALL_OPPS_COLUMNS
from src.storage.database import session_scope
from src.storage.models import AuditLog, Opportunity, Run, User
from src.storage.repository import OpportunityRepo
from src.web.auth import (
    COOKIE_NAME, CurrentUser, authenticate, check_rate_limit,
    create_token, get_current_user, hash_password, record_login_attempt,
    require_admin,
)

logger = get_logger(__name__)

TEMPLATES_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

router = APIRouter()


# ============================================================
# Helper : audit
# ============================================================

async def _audit(
    session,
    user: CurrentUser | None,
    action: str,
    request: Request,
    resource_type: str | None = None,
    resource_id: str | None = None,
    details: dict | None = None,
) -> None:
    log = AuditLog(
        user_id=user.id if user else None,
        user_email=user.email if user else None,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        ip_address=request.client.host if request.client else None,
        details=details,
    )
    session.add(log)


# ============================================================
# Pages publiques
# ============================================================

@router.get("/", response_class=HTMLResponse)
async def root(request: Request, session_cookie: Annotated[str | None, Cookie(alias=COOKIE_NAME)] = None):
    if session_cookie:
        return RedirectResponse(url="/dashboard", status_code=302)
    return RedirectResponse(url="/login", status_code=302)


@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str | None = None):
    return templates.TemplateResponse(
            request=request,
            name="login.html",
            context={"request": request, "error": error}
        )


@router.post("/login")
async def login_submit(
    request: Request,
    response: Response,
    email: Annotated[str, Form()],
    password: Annotated[str, Form()],
):
    ip = request.client.host if request.client else "unknown"

    if not check_rate_limit(ip):
        logger.warning("rate_limit_exceeded", ip=ip)
        return templates.TemplateResponse(
            request=request,
            name="login.html",
            context={
                "request": request,
                "error": "Trop de tentatives. Réessayez dans 15 minutes.",
            },
            status_code=429
        )

    record_login_attempt(ip)

    async with session_scope() as session:
        user = await authenticate(email, password, session)
        if user is None:
            await _audit(session, None, "LOGIN_FAIL", request, details={"email": email})
            return templates.TemplateResponse(
            request=request,
            name="login.html",
            context={"request": request, "error": "Email ou mot de passe incorrect."},
            status_code=401
        )

        token = create_token(user.id, user.email, user.role)
        await _audit(
            session,
            CurrentUser(id=user.id, email=user.email, role=user.role),
            "LOGIN_OK",
            request,
        )

    response = RedirectResponse(url="/dashboard", status_code=302)
    # Cookie sécurisé : httpOnly (anti-XSS), secure (HTTPS only), sameSite=strict (anti-CSRF)
    response.set_cookie(
        key=COOKIE_NAME,
        value=token,
        max_age=8 * 3600,
        httponly=True,
        secure=True,
        samesite="strict",
        path="/",
    )
    return response


@router.post("/logout")
async def logout(request: Request, user: Annotated[CurrentUser, Depends(get_current_user)]):
    async with session_scope() as session:
        await _audit(session, user, "LOGOUT", request)
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie(COOKIE_NAME, path="/")
    return response


# ============================================================
# Dashboard
# ============================================================

@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    days: int = Query(default=7, ge=1, le=90),
    min_score: int = Query(default=0, ge=0, le=100),
):
    async with session_scope() as session:
        repo = OpportunityRepo(session)
        opps = await repo.list_recent(days=days, only_open=True, min_score=min_score)

        # Stats globales
        total = len(opps)
        high = sum(1 for o in opps if (o.score or 0) >= 70)
        qualified = sum(1 for o in opps if (o.score or 0) >= 50)
        urgent = sum(1 for o in opps if o.is_urgent)

        # Décisions déjà prises ce cycle
        decided = sum(1 for o in opps if o.client_decision)
        pending = total - decided

        # Dernier run
        last_run_stmt = select(Run).order_by(desc(Run.started_at)).limit(1)
        last_run = (await session.execute(last_run_stmt)).scalar_one_or_none()

    return templates.TemplateResponse(
            request=request,
            name="dashboard.html",
            context={
            "request": request,
            "user": user,
            "opportunities": opps,
            "stats": {
                "total": total,
                "high": high,
                "qualified": qualified,
                "urgent": urgent,
                "decided": decided,
                "pending": pending,
            },
            "filter_days": days,
            "filter_min_score": min_score,
            "last_run": last_run,
        }
        )


@router.get("/opportunity/{opp_id}", response_class=HTMLResponse)
async def opportunity_detail(
    request: Request,
    opp_id: int,
    user: Annotated[CurrentUser, Depends(get_current_user)],
):
    async with session_scope() as session:
        opp = (await session.execute(
            select(Opportunity).where(Opportunity.id == opp_id)
        )).scalar_one_or_none()
        if opp is None:
            raise HTTPException(404, "Opportunité introuvable")

    return templates.TemplateResponse(
            request=request,
            name="opportunity.html",
            context={"request": request, "user": user, "opp": opp}
        )


# ============================================================
# API : décisions (1-clic depuis dashboard)
# ============================================================

@router.post("/api/decision")
async def api_decision(
    request: Request,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    opportunity_id: Annotated[int, Form()],
    decision: Annotated[str, Form()],
    reason: Annotated[str | None, Form()] = None,
):
    decision = decision.upper().strip()
    if decision not in {"GO", "NO_GO", "BORDERLINE", "SUBMITTED", ""}:
        raise HTTPException(400, "Décision invalide")

    async with session_scope() as session:
        if decision == "":
            # Reset
            opp = (await session.execute(
                select(Opportunity).where(Opportunity.id == opportunity_id)
            )).scalar_one()
            opp.client_decision = None
            opp.client_reason = None
            opp.client_decided_at = None
            opp.status = "DETECTED"
            action = "DECISION_RESET"
        else:
            repo = OpportunityRepo(session)
            await repo.apply_feedback(
                opportunity_id=opportunity_id,
                decision=decision,
                reason=reason,
                submitted_by=user.email,
            )
            action = f"DECISION_{decision}"

        await _audit(
            session, user, action, request,
            resource_type="opportunity",
            resource_id=str(opportunity_id),
            details={"reason": reason} if reason else None,
        )

    return {"ok": True, "decision": decision or "RESET"}


# ============================================================
# API : upload Excel
# ============================================================

COL_ID = ALL_OPPS_COLUMNS.index("ID")
COL_TITLE = ALL_OPPS_COLUMNS.index("Titre")
COL_DECISION = ALL_OPPS_COLUMNS.index("Décision interne (Go/No-Go)")
COL_REASON = ALL_OPPS_COLUMNS.index("Raison décision")
VALID_DECISIONS = {"GO", "NO_GO", "BORDERLINE", "SUBMITTED"}


@router.post("/api/upload-excel")
async def api_upload_excel(
    request: Request,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    file: Annotated[UploadFile, File()],
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Fichier .xlsx requis")

    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(413, "Fichier > 20 MB rejeté")

    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel illisible : {e}")

    sheet = "Toutes opportunités"
    if sheet not in wb.sheetnames:
        raise HTTPException(400, f"Onglet '{sheet}' absent")
    ws = wb[sheet]

    n_processed, n_skipped, errors = 0, 0, []
    async with session_scope() as session:
        repo = OpportunityRepo(session)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[COL_ID] is None:
                continue
            decision_raw = row[COL_DECISION]
            if not decision_raw:
                n_skipped += 1
                continue
            decision = str(decision_raw).strip().upper()
            if decision not in VALID_DECISIONS:
                errors.append({"row": i, "error": f"Décision invalide '{decision}'"})
                continue
            try:
                await repo.apply_feedback(
                    opportunity_id=int(row[COL_ID]),
                    decision=decision,
                    reason=str(row[COL_REASON]) if row[COL_REASON] else None,
                    submitted_by=user.email,
                )
                n_processed += 1
            except Exception as e:
                errors.append({"row": i, "error": str(e)[:200]})

        await _audit(
            session, user, "EXCEL_UPLOAD", request,
            details={"processed": n_processed, "errors": len(errors)},
        )

    return {
        "ok": True, "processed": n_processed,
        "skipped": n_skipped, "errors": errors,
    }


# ============================================================
# Admin : utilisateurs
# ============================================================

@router.get("/admin/users", response_class=HTMLResponse)
async def admin_users(
    request: Request,
    user: Annotated[CurrentUser, Depends(require_admin)],
):
    async with session_scope() as session:
        users = (await session.execute(
            select(User).order_by(User.created_at.desc())
        )).scalars().all()
    return templates.TemplateResponse(
            request=request,
            name="admin_users.html",
            context={"request": request, "user": user, "users": list(users)}
        )


@router.post("/admin/users")
async def admin_create_user(
    request: Request,
    user: Annotated[CurrentUser, Depends(require_admin)],
    email: Annotated[str, Form()],
    full_name: Annotated[str, Form()],
    password: Annotated[str, Form()],
    role: Annotated[str, Form()] = "analyst",
):
    if len(password) < 12:
        raise HTTPException(400, "Mot de passe : 12 caractères minimum")
    if role not in {"admin", "analyst", "viewer"}:
        raise HTTPException(400, "Rôle invalide")

    async with session_scope() as session:
        existing = (await session.execute(
            select(User).where(User.email == email.lower())
        )).scalar_one_or_none()
        if existing:
            raise HTTPException(409, "Email déjà utilisé")

        new_user = User(
            email=email.lower().strip(),
            full_name=full_name.strip(),
            password_hash=hash_password(password),
            role=role,
            is_active=True,
        )
        session.add(new_user)
        await session.flush()
        await _audit(
            session, user, "USER_CREATED", request,
            resource_type="user", resource_id=str(new_user.id),
            details={"email": email, "role": role},
        )

    return RedirectResponse(url="/admin/users", status_code=302)


# ============================================================
# Admin : historique des runs
# ============================================================

@router.get("/admin/runs", response_class=HTMLResponse)
async def admin_runs(
    request: Request,
    user: Annotated[CurrentUser, Depends(require_admin)],
):
    async with session_scope() as session:
        runs = (await session.execute(
            select(Run).order_by(Run.started_at.desc()).limit(50)
        )).scalars().all()

        # Stats globales
        total_opps = (await session.execute(
            select(func.count(Opportunity.id))
        )).scalar()
        decided = (await session.execute(
            select(func.count(Opportunity.id)).where(Opportunity.client_decision.isnot(None))
        )).scalar()

        # 5 derniers événements d'audit
        audits = (await session.execute(
            select(AuditLog).order_by(AuditLog.timestamp.desc()).limit(20)
        )).scalars().all()

    return templates.TemplateResponse(
            request=request,
            name="admin_runs.html",
            context={
            "request": request, "user": user,
            "runs": list(runs),
            "audits": list(audits),
            "stats": {"total_opps": total_opps, "decided": decided},
        }
        )


@router.post("/admin/trigger-run")
async def admin_trigger_run(
    request: Request,
    user: Annotated[CurrentUser, Depends(require_admin)],
):
    """Déclenche un run hebdo manuel (pour test/urgence)."""
    import asyncio
    from scripts.run_weekly import main as run_weekly

    async with session_scope() as session:
        await _audit(session, user, "RUN_TRIGGERED", request)

    # Lance en background, sans bloquer la réponse
    asyncio.create_task(run_weekly())

    return {"ok": True, "msg": "Run lancé en arrière-plan. Voir /admin/runs dans quelques minutes."}
