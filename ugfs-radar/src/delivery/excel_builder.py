"""
src/delivery/excel_builder.py — Fichier Excel hebdomadaire UGFS.

QUALITÉ CIBLE : cabinet de conseil tier 1.
On livre un fichier en .xlsx avec :

  Onglet 1 — Dashboard      Synthèse semaine, KPI, top 5 opportunités
  Onglet 2 — Toutes opps    Tableau exhaustif trié par score + colonne FEEDBACK
  Onglet 3 — Fiches détail  Une fiche structurée par opportunité
  Onglet 4 — Historique     Comparaison avec les soumissions passées

DESIGN
------
- Charte UGFS : bleu marine + accents verts (cohérent avec PE/Green/Blue funds)
- Conditional formatting sur les scores (rouge < 50, jaune 50-69, vert ≥ 70)
- Colonnes figées, filtres auto, dropdowns Go/No-Go pré-configurés
- Pas de macros (sécurité IT entreprise)

FEEDBACK MECHANISM
------------------
La colonne `Décision interne` de l'onglet 2 est une dropdown EXCEL native
(data validation). Le client la remplit après sa réunion interne hebdo,
renvoie le fichier, et `scripts/ingest_feedback.py` lit ce même fichier.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

from src.config.logger import get_logger
from src.storage.models import Opportunity

logger = get_logger(__name__)


# ============================================================
# Charte graphique UGFS
# ============================================================

class Theme:
    # Bleu marine UGFS (couleur principale)
    PRIMARY = "0F2A4A"
    PRIMARY_LIGHT = "D6E2EE"

    # Vert (Tunisia Green Fund)
    GREEN_DARK = "1F6B3A"
    GREEN_LIGHT = "D8EBDD"

    # Bleu (Blue Bond)
    BLUE_DARK = "1B4F8C"
    BLUE_LIGHT = "D5E4F2"

    # Scoring
    SCORE_HIGH = "C6EFCE"          # Vert pâle
    SCORE_MID = "FFEB9C"           # Jaune pâle
    SCORE_LOW = "F8CBAD"           # Orange pâle
    SCORE_DQ = "F4B084"            # Orange plus foncé

    # Texte
    TEXT_PRIMARY = "FFFFFF"        # blanc sur primary
    TEXT_DARK = "1A1A1A"
    TEXT_MUTED = "6B7280"

    # Lignes / bordures
    BORDER = "CBD5E1"


# ============================================================
# Styles réutilisables
# ============================================================

THIN_SIDE = Side(style="thin", color=Theme.BORDER)
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=Theme.PRIMARY)
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=Theme.TEXT_MUTED)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=Theme.TEXT_PRIMARY)
FONT_BODY = Font(name="Calibri", size=10, color=Theme.TEXT_DARK)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=Theme.TEXT_DARK)
FONT_KPI_LABEL = Font(name="Calibri", size=10, color=Theme.TEXT_MUTED)
FONT_KPI_VALUE = Font(name="Calibri", size=22, bold=True, color=Theme.PRIMARY)

FILL_HEADER = PatternFill(start_color=Theme.PRIMARY, end_color=Theme.PRIMARY, fill_type="solid")
FILL_KPI_BOX = PatternFill(start_color=Theme.PRIMARY_LIGHT, end_color=Theme.PRIMARY_LIGHT, fill_type="solid")
FILL_GREEN = PatternFill(start_color=Theme.GREEN_LIGHT, end_color=Theme.GREEN_LIGHT, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


# ============================================================
# Helper : style d'en-tête
# ============================================================

def _style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 30


def _apply_table_borders(ws, start_row: int, end_row: int, cols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL


def _apply_score_cf(ws, range_str: str) -> None:
    """Conditional formatting sur le score : vert / jaune / orange / rouge."""
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="greaterThanOrEqual", formula=["70"],
                   fill=PatternFill(start_color=Theme.SCORE_HIGH, end_color=Theme.SCORE_HIGH, fill_type="solid")),
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="between", formula=["50", "69"],
                   fill=PatternFill(start_color=Theme.SCORE_MID, end_color=Theme.SCORE_MID, fill_type="solid")),
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="between", formula=["1", "49"],
                   fill=PatternFill(start_color=Theme.SCORE_LOW, end_color=Theme.SCORE_LOW, fill_type="solid")),
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="equal", formula=["0"],
                   fill=PatternFill(start_color=Theme.SCORE_DQ, end_color=Theme.SCORE_DQ, fill_type="solid")),
    )


# ============================================================
# Onglet 1 — DASHBOARD
# ============================================================

def _build_dashboard(ws, opportunities: Sequence[Opportunity], run_date: date) -> None:
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # En-tête
    ws.merge_cells("A1:H1")
    ws["A1"] = "UGFS-Radar — Synthèse hebdomadaire"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Édition du {run_date.strftime('%d/%m/%Y')} · "
        f"Veille automatisée d'opportunités d'investissement et de financement"
    )
    ws["A2"].font = FONT_SUBTITLE

    # === KPI ===
    total = len(opportunities)
    qualified = sum(1 for o in opportunities if (o.score or 0) >= 50)
    high_score = sum(1 for o in opportunities if (o.score or 0) >= 70)
    urgent = sum(1 for o in opportunities if o.is_urgent)

    kpis = [
        ("Opportunités détectées", total),
        ("Qualifiées (score ≥ 50)", qualified),
        ("Prioritaires (score ≥ 70)", high_score),
        ("URGENTES (≤ 7 jours)", urgent),
    ]
    start_row = 4
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 2  # A, C, E, G
        # Box
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=start_row, end_column=col + 1)
        ws.cell(row=start_row, column=col, value=label)
        ws.cell(row=start_row, column=col).font = FONT_KPI_LABEL
        ws.cell(row=start_row, column=col).alignment = ALIGN_CENTER
        ws.cell(row=start_row, column=col).fill = FILL_KPI_BOX

        ws.merge_cells(start_row=start_row + 1, start_column=col,
                       end_row=start_row + 2, end_column=col + 1)
        ws.cell(row=start_row + 1, column=col, value=value)
        ws.cell(row=start_row + 1, column=col).font = FONT_KPI_VALUE
        ws.cell(row=start_row + 1, column=col).alignment = ALIGN_CENTER
        ws.cell(row=start_row + 1, column=col).fill = FILL_KPI_BOX

    ws.row_dimensions[start_row].height = 22
    ws.row_dimensions[start_row + 1].height = 30
    ws.row_dimensions[start_row + 2].height = 18

    # === TOP 5 ===
    title_row = start_row + 4
    ws.cell(row=title_row, column=1, value="🏆 Top 5 opportunités de la semaine")
    ws.cell(row=title_row, column=1).font = Font(name="Calibri", size=14, bold=True, color=Theme.PRIMARY)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)

    headers = ["Rang", "Score", "Titre", "Type", "Géo", "Deadline", "Véhicule UGFS", "Source"]
    hdr_row = title_row + 1
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    _style_header_row(ws, hdr_row, len(headers))

    top = sorted(opportunities, key=lambda o: o.score or 0, reverse=True)[:5]
    for rank, opp in enumerate(top, start=1):
        r = hdr_row + rank
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=opp.score)
        ws.cell(row=r, column=3, value=opp.title)
        ws.cell(row=r, column=4, value=opp.opportunity_type or "—")
        ws.cell(row=r, column=5, value=", ".join(opp.geographies or [])[:60] or "—")
        ws.cell(row=r, column=6,
                value=opp.deadline.strftime("%d/%m/%Y") if opp.deadline else (opp.deadline_text_raw or "Rolling"))
        ws.cell(row=r, column=7, value=opp.vehicle_match or "—")
        ws.cell(row=r, column=8, value=opp.source or "—")
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).alignment = ALIGN_LEFT_WRAP
            ws.cell(row=r, column=col).font = FONT_BODY
            ws.cell(row=r, column=col).border = BORDER_ALL
        ws.row_dimensions[r].height = 30

    if top:
        _apply_score_cf(ws, f"B{hdr_row + 1}:B{hdr_row + len(top)}")

    # Largeurs
    widths = [6, 8, 50, 14, 25, 14, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer note
    foot = hdr_row + len(top) + 3
    ws.cell(row=foot, column=1, value="ℹ️ Le détail complet est dans l'onglet « Toutes opportunités ».")
    ws.cell(row=foot, column=1).font = FONT_SUBTITLE
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=8)


# ============================================================
# Onglet 2 — TOUTES OPPORTUNITÉS (avec colonne FEEDBACK)
# ============================================================

# IMPORTANT : ces colonnes sont lues par scripts/ingest_feedback.py.
# Toute modification doit être reflétée dans `INGEST_COLUMN_MAP`.
ALL_OPPS_COLUMNS = [
    "ID", "Score", "Titre", "Type", "Thème", "Géographies",
    "Deadline", "Ticket (USD)", "Véhicule UGFS", "Partenaires",
    "Statut", "URL", "Décision interne (Go/No-Go)", "Raison décision",
]


def _build_all_opportunities(ws, opportunities: Sequence[Opportunity]) -> None:
    ws.title = "Toutes opportunités"
    ws.sheet_view.showGridLines = False

    # Header
    for i, h in enumerate(ALL_OPPS_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(ALL_OPPS_COLUMNS))
    ws.freeze_panes = "A2"

    # Données triées par score
    sorted_opps = sorted(opportunities, key=lambda o: o.score or 0, reverse=True)
    for i, opp in enumerate(sorted_opps, start=2):
        ws.cell(row=i, column=1, value=opp.id)
        ws.cell(row=i, column=2, value=opp.score)
        ws.cell(row=i, column=3, value=opp.title)
        ws.cell(row=i, column=4, value=opp.opportunity_type or "—")
        ws.cell(row=i, column=5, value=opp.theme or "—")
        ws.cell(row=i, column=6, value=", ".join(opp.geographies or []) or "—")
        ws.cell(row=i, column=7,
                value=opp.deadline.strftime("%d/%m/%Y") if opp.deadline else (opp.deadline_text_raw or "Rolling"))
        ws.cell(row=i, column=8,
                value=f"{opp.ticket_size_usd:,}" if opp.ticket_size_usd else "—")
        ws.cell(row=i, column=9, value=opp.vehicle_match or "—")
        ws.cell(row=i, column=10, value=", ".join((opp.partners_mentioned or []))[:80] or "—")
        ws.cell(row=i, column=11, value=opp.status or "DETECTED")
        # Hyperlien vers la source
        url_cell = ws.cell(row=i, column=12, value=opp.url)
        url_cell.hyperlink = opp.url
        url_cell.font = Font(name="Calibri", size=10, color="1B4F8C", underline="single")
        # Colonnes feedback (vide, à remplir par UGFS)
        ws.cell(row=i, column=13, value=opp.client_decision or "")
        ws.cell(row=i, column=14, value=opp.client_reason or "")

        for c in range(1, len(ALL_OPPS_COLUMNS) + 1):
            cell = ws.cell(row=i, column=c)
            cell.alignment = ALIGN_LEFT_WRAP
            cell.font = FONT_BODY
            cell.border = BORDER_ALL
        # Colonne décision en bold pour attirer l'œil
        ws.cell(row=i, column=13).font = FONT_BODY_BOLD
        ws.cell(row=i, column=13).fill = FILL_GREEN

        ws.row_dimensions[i].height = 36

    # Filtres auto + freeze
    if sorted_opps:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_OPPS_COLUMNS))}{1 + len(sorted_opps)}"
        _apply_score_cf(ws, f"B2:B{1 + len(sorted_opps)}")

        # Data validation : dropdown Go/No-Go/Borderline sur la colonne 13
        dv = DataValidation(
            type="list",
            formula1='"GO,NO_GO,BORDERLINE,SUBMITTED"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Décision invalide",
            error="Choisir une valeur dans la liste : GO / NO_GO / BORDERLINE / SUBMITTED",
        )
        ws.add_data_validation(dv)
        dv.add(f"M2:M{1 + len(sorted_opps)}")

    # Largeurs
    widths = [6, 8, 50, 14, 12, 26, 14, 14, 18, 26, 12, 30, 22, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Onglet 3 — FICHES DÉTAILLÉES
# ============================================================

def _build_detail_cards(ws, opportunities: Sequence[Opportunity]) -> None:
    ws.title = "Fiches détaillées"
    ws.sheet_view.showGridLines = False

    row = 1
    for opp in sorted(opportunities, key=lambda o: o.score or 0, reverse=True):
        # Titre carte
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1,
                value=f"#{opp.id} · Score {opp.score}/100 · {opp.title}")
        ws.cell(row=row, column=1).font = Font(
            name="Calibri", size=13, bold=True, color=Theme.TEXT_PRIMARY,
        )
        ws.cell(row=row, column=1).fill = FILL_HEADER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1

        # Tableau de propriétés
        rows_data = [
            ("Type", opp.opportunity_type or "—"),
            ("Thème", opp.theme or "—"),
            ("Géographies", ", ".join(opp.geographies or []) or "—"),
            ("Secteurs", ", ".join(opp.sectors or []) or "—"),
            ("Deadline",
             opp.deadline.strftime("%d/%m/%Y")
             if opp.deadline else (opp.deadline_text_raw or "Rolling")),
            ("Ticket size", f"{opp.ticket_size_usd:,} USD" if opp.ticket_size_usd else "—"),
            ("Langues", ", ".join(opp.languages or []) or "—"),
            ("Véhicule UGFS matché", opp.vehicle_match or "Aucun"),
            ("Partenaires mentionnés", ", ".join(opp.partners_mentioned or []) or "—"),
            ("URL source", opp.url or "—"),
            ("URL formulaire soumission", opp.submission_url or "—"),
            ("Statut", opp.status or "DETECTED"),
            ("Source", opp.source or "—"),
        ]
        for label, value in rows_data:
            ws.cell(row=row, column=1, value=label).font = FONT_BODY_BOLD
            ws.cell(row=row, column=1).fill = FILL_KPI_BOX
            ws.cell(row=row, column=1).alignment = ALIGN_LEFT_WRAP
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row=row, column=2, value=str(value)).font = FONT_BODY
            ws.cell(row=row, column=2).alignment = ALIGN_LEFT_WRAP
            for c in (1, 2):
                ws.cell(row=row, column=c).border = BORDER_ALL
            row += 1

        # Sections texte (Résumé exécutif, Pourquoi, Éligibilité)
        text_sections = [
            ("📋 Résumé exécutif", opp.summary_executive or "—"),
            ("✅ Pourquoi c'est intéressant", opp.why_interesting or "—"),
            ("📜 Éligibilité", opp.eligibility_summary or "—"),
            ("🤖 Recommandation préliminaire IA",
             f"{opp.preliminary_decision or 'PENDING'} — {opp.decision_rationale or '—'}"),
        ]
        for header, text in text_sections:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=header).font = FONT_BODY_BOLD
            ws.cell(row=row, column=1).fill = FILL_GREEN
            ws.cell(row=row, column=1).alignment = ALIGN_LEFT_WRAP
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=text).font = FONT_BODY
            ws.cell(row=row, column=1).alignment = ALIGN_LEFT_WRAP
            ws.row_dimensions[row].height = max(40, min(120, 15 * (1 + len(text) // 90)))
            row += 1

        # Saut de carte
        row += 2

    widths = [22, 30, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Onglet 4 — HISTORIQUE
# ============================================================

def _build_history(ws, historical: Sequence[Opportunity]) -> None:
    ws.title = "Historique"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Comparaison avec les soumissions UGFS passées"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")

    ws["A2"] = (
        "Cet onglet liste les opportunités historiques renseignées par UGFS "
        "(seed initial + décisions client accumulées). C'est la base de calibration "
        "du scoring : les nouveaux AOs proches de Go passés sont automatiquement boostés."
    )
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_LEFT_WRAP
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 30

    headers = ["Titre", "Type", "Thème", "Géo", "Deadline", "Décision", "Raison"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))
    ws.freeze_panes = "A5"

    for i, opp in enumerate(historical, start=5):
        ws.cell(row=i, column=1, value=opp.title)
        ws.cell(row=i, column=2, value=opp.opportunity_type or "—")
        ws.cell(row=i, column=3, value=opp.theme or "—")
        ws.cell(row=i, column=4, value=", ".join(opp.geographies or []) or "—")
        ws.cell(row=i, column=5,
                value=opp.deadline.strftime("%d/%m/%Y") if opp.deadline else "—")
        ws.cell(row=i, column=6, value=opp.client_decision or "UNKNOWN")
        ws.cell(row=i, column=7, value=(opp.client_reason or "")[:200])
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).font = FONT_BODY
            ws.cell(row=i, column=c).alignment = ALIGN_LEFT_WRAP
            ws.cell(row=i, column=c).border = BORDER_ALL
        ws.row_dimensions[i].height = 30

    if historical:
        ws.auto_filter.ref = f"A4:G{4 + len(historical)}"

    widths = [50, 18, 14, 26, 14, 14, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# API publique
# ============================================================

def build_weekly_excel(
    opportunities: Sequence[Opportunity],
    historical: Sequence[Opportunity],
    output_path: str | Path | None = None,
    run_date: date | None = None,
) -> bytes:
    """
    Construit le fichier Excel hebdomadaire UGFS.

    Args:
        opportunities: opportunités scorées de la semaine
        historical: opportunités historiques (avec décisions)
        output_path: si fourni, sauvegarde sur disque
        run_date: date d'édition (par défaut aujourd'hui)

    Returns:
        Le fichier en bytes (pour envoi email direct).
    """
    if run_date is None:
        run_date = date.today()

    wb = Workbook()
    # Onglet 1 (par défaut)
    ws_dash = wb.active
    _build_dashboard(ws_dash, opportunities, run_date)

    # Onglet 2
    ws_all = wb.create_sheet()
    _build_all_opportunities(ws_all, opportunities)

    # Onglet 3
    ws_cards = wb.create_sheet()
    _build_detail_cards(ws_cards, opportunities)

    # Onglet 4
    ws_hist = wb.create_sheet()
    _build_history(ws_hist, historical)

    # Métadonnées workbook
    wb.properties.title = "UGFS-Radar — Veille hebdomadaire"
    wb.properties.subject = f"Édition {run_date.strftime('%d/%m/%Y')}"
    wb.properties.creator = "UGFS-Radar Agent"
    wb.properties.company = "UGFS North Africa"

    # Sauvegarde
    buffer = io.BytesIO()
    wb.save(buffer)
    data = buffer.getvalue()

    if output_path:
        Path(output_path).write_bytes(data)
        logger.info(
            "excel_built",
            path=str(output_path),
            n_opps=len(opportunities),
            n_historical=len(historical),
            size_kb=round(len(data) / 1024, 1),
        )

    return data
