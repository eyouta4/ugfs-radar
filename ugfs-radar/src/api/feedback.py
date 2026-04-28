"""
src/api/feedback.py — Ingestion des décisions Go/No-Go d'UGFS.

Le client renvoie chaque semaine le fichier Excel avec sa colonne
« Décision interne (Go/No-Go) » remplie. Cet endpoint :
  1. Reçoit le fichier en upload (multipart/form-data)
  2. Lit l'onglet « Toutes opportunités »
  3. Pour chaque ligne avec décision non-vide → applique le feedback en DB
  4. Retourne un résumé (N feedbacks ingérés, erreurs éventuelles)

Le déclenchement de la recalibration des poids se fait séparément via
`scripts/recalibrate_weights.py` (ou via un endpoint admin futur).
"""
from __future__ import annotations

import io
from typing import Annotated

from fastapi import APIRouter, File, Header, HTTPException, UploadFile
from openpyxl import load_workbook

from src.config.logger import get_logger
from src.config.settings import get_settings
from src.delivery.excel_builder import ALL_OPPS_COLUMNS
from src.storage.database import session_scope
from src.storage.repository import OpportunityRepo

router = APIRouter()
logger = get_logger(__name__)

VALID_DECISIONS = {"GO", "NO_GO", "BORDERLINE", "SUBMITTED"}

# Map nom de colonne → index (basé sur ALL_OPPS_COLUMNS de excel_builder.py)
COL_ID = 0  # Opportunite (colonne A)
COL_TITLE = 0  # Opportunite
COL_DECISION = 7  # Responsable / Decision UGFS (colonne H)
COL_REASON = 7  # meme colonne


def _check_token(token: str | None) -> None:
    settings = get_settings()
    if not token or token != settings.api_feedback_token:
        raise HTTPException(status_code=401, detail="Invalid or missing X-API-Token")


@router.post("/excel")
async def ingest_excel_feedback(
    file: Annotated[UploadFile, File(description="Fichier Excel UGFS-Radar avec décisions")],
    x_api_token: Annotated[str | None, Header(alias="X-API-Token")] = None,
    submitted_by: str | None = None,
) -> dict:
    """
    Ingère les décisions Go/No-Go renseignées dans le fichier Excel hebdo.
    """
    _check_token(x_api_token)

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Fichier .xlsx requis")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel illisible : {e}")

    sheet_name = "Toutes opportunités"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"Onglet '{sheet_name}' introuvable. Onglets: {wb.sheetnames}",
        )
    ws = wb[sheet_name]

    n_processed = 0
    n_skipped = 0
    errors: list[dict] = []

    async with session_scope() as session:
        repo = OpportunityRepo(session)

        # Skip header row
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[COL_ID]:
                continue

            opp_id_raw = row[COL_ID]
            decision = (str(row[COL_DECISION]) or "").strip().upper() if row[COL_DECISION] else ""
            reason = str(row[COL_REASON] or "").strip() if row[COL_REASON] else None
            title = str(row[COL_TITLE] or "")[:80] if row[COL_TITLE] else "?"

            if not decision:
                n_skipped += 1
                continue

            if decision not in VALID_DECISIONS:
                errors.append({
                    "row": i,
                    "id": opp_id_raw,
                    "title": title,
                    "error": f"Décision invalide : '{decision}'. Attendu : {VALID_DECISIONS}",
                })
                continue

            try:
                opp_id = int(opp_id_raw)
            except (TypeError, ValueError):
                errors.append({
                    "row": i,
                    "id": opp_id_raw,
                    "title": title,
                    "error": "ID non-numérique",
                })
                continue

            try:
                await repo.apply_feedback(
                    opportunity_id=opp_id,
                    decision=decision,
                    reason=reason,
                    submitted_by=submitted_by or "excel_upload",
                )
                n_processed += 1
            except Exception as e:
                errors.append({
                    "row": i, "id": opp_id, "title": title, "error": str(e)[:200],
                })

    logger.info(
        "feedback_ingested",
        processed=n_processed,
        skipped=n_skipped,
        errors=len(errors),
    )

    return {
        "ok": True,
        "processed": n_processed,
        "skipped_empty": n_skipped,
        "errors": errors,
    }


@router.post("/single")
async def ingest_single_feedback(
    opportunity_id: int,
    decision: str,
    reason: str | None = None,
    submitted_by: str | None = None,
    x_api_token: Annotated[str | None, Header(alias="X-API-Token")] = None,
) -> dict:
    """Ingestion d'une seule décision (utile pour intégrations futures)."""
    _check_token(x_api_token)
    decision = decision.strip().upper()
    if decision not in VALID_DECISIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Décision invalide. Attendu : {VALID_DECISIONS}",
        )
    async with session_scope() as session:
        repo = OpportunityRepo(session)
        await repo.apply_feedback(
            opportunity_id=opportunity_id,
            decision=decision,
            reason=reason,
            submitted_by=submitted_by,
        )
    return {"ok": True, "opportunity_id": opportunity_id, "decision": decision}
