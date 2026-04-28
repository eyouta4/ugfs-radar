"""tests/test_excel_builder.py — Tests de génération du fichier Excel."""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from types import SimpleNamespace

from openpyxl import load_workbook

from src.delivery.excel_builder import ALL_OPPS_COLUMNS, build_weekly_excel


def _fake_opp(
    id_: int,
    title: str,
    score: int = 75,
    is_urgent: bool = False,
    deadline_days: int | None = 30,
    decision: str | None = None,
):
    return SimpleNamespace(
        id=id_,
        title=title,
        score=score,
        url=f"https://example.org/ao/{id_}",
        source="test",
        opportunity_type="grant",
        theme="green",
        geographies=["Tunisia"],
        sectors=["climate"],
        deadline=date.today() + timedelta(days=deadline_days) if deadline_days else None,
        deadline_text_raw="Rolling" if deadline_days is None else None,
        ticket_size_usd=1_000_000,
        languages=["fr", "en"],
        vehicle_match="TGF",
        partners_mentioned=["GIZ"],
        status="DETECTED",
        is_urgent=is_urgent,
        summary_executive="Résumé exécutif test.",
        why_interesting="Match parfait UGFS.",
        eligibility_summary="Asset managers éligibles.",
        preliminary_decision="GO",
        decision_rationale="Tout aligné.",
        submission_url=None,
        client_decision=decision,
        client_reason=None,
    )


def test_excel_builds_without_error():
    opps = [_fake_opp(1, "Climate Fund A"), _fake_opp(2, "Blue Bond B", score=55)]
    historical = [_fake_opp(99, "Past Go", decision="GO_SUBMITTED")]
    data = build_weekly_excel(opps, historical, run_date=date(2026, 4, 27))
    assert len(data) > 5000  # un fichier xlsx avec contenu non-trivial


def test_excel_has_four_sheets():
    opps = [_fake_opp(1, "Climate Fund A")]
    data = build_weekly_excel(opps, [], run_date=date(2026, 4, 27))
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == [
        "Dashboard", "Toutes opportunités", "Fiches détaillées", "Historique",
    ]


def test_excel_dashboard_kpis():
    opps = [
        _fake_opp(1, "A", score=80),
        _fake_opp(2, "B", score=65),
        _fake_opp(3, "C", score=30),
        _fake_opp(4, "D", score=90, is_urgent=True, deadline_days=5),
    ]
    data = build_weekly_excel(opps, [], run_date=date(2026, 4, 27))
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Dashboard"]

    # Title cell
    assert "UGFS-Radar" in str(ws["A1"].value)


def test_excel_all_opps_has_correct_columns():
    opps = [_fake_opp(1, "A")]
    data = build_weekly_excel(opps, [], run_date=date(2026, 4, 27))
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Toutes opportunités"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, len(ALL_OPPS_COLUMNS) + 1)]
    assert headers == ALL_OPPS_COLUMNS


def test_excel_handles_empty_opportunities():
    """Le build doit fonctionner même sans opportunités (semaine creuse)."""
    data = build_weekly_excel([], [], run_date=date(2026, 4, 27))
    wb = load_workbook(io.BytesIO(data))
    assert "Dashboard" in wb.sheetnames
